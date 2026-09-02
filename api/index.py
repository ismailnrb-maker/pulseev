import os
import csv
import io
import uuid
import datetime
import re
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session

from api.database import get_db, init_db, User, Vehicle, ActionCase, ActionAuditEvent, ServiceCampaign
from api.auth import hash_password, verify_password, create_access_token, get_current_user
from api.action_centre import draft_whatsapp, enrich_action_centre, iso_now, serialize_case, sync_cases

# Initialize app
app = FastAPI(title="PulseEV API", docs_url="/api/docs", openapi_url="/api/openapi.json")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Startup DB init — fault-tolerant so a DB error doesn't crash all endpoints
@app.on_event("startup")
def startup_event():
    try:
        init_db()
    except Exception as e:
        print(f"startup init_db error (non-fatal): {e}")

# ── Health / Debug endpoint ───────────────────────────────────────────────────
@app.get("/api/health")
def health():
    import os
    from api.database import DATABASE_URL, IS_POSTGRES
    db_type = "postgres" if IS_POSTGRES else "sqlite"
    masked = DATABASE_URL[:40] + "..." if len(DATABASE_URL) > 40 else DATABASE_URL
    try:
        from api.database import engine
        with engine.connect() as conn:
            conn.execute(__import__('sqlalchemy').text('SELECT 1'))
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)[:200]}"
    return {"status": "ok", "db_type": db_type, "db_url_prefix": masked, "db_status": db_status}

# --- Schemas ---
from pydantic import BaseModel, Field

class LoginRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str


class ActionRequest(BaseModel):
    type: str
    payload: dict = Field(default_factory=dict)


class EnrichRequest(BaseModel):
    force: bool = False


def validate_vehicle_identity(payload: dict, partial: bool = False):
    """Reject pitch-breaking identity and contact data at the API boundary."""
    vin = str(payload.get("vin", "")).strip().upper()
    name = str(payload.get("customerName", "")).strip()
    phone = str(payload.get("customerPhone", "")).strip()
    delivery = str(payload.get("deliveryDate", "")).strip()
    if (not partial or "vin" in payload) and not re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", vin):
        raise HTTPException(status_code=400, detail="VIN must contain exactly 17 valid characters")
    if (not partial or "customerName" in payload) and (len(name) < 5 or " " not in name):
        raise HTTPException(status_code=400, detail="Enter the customer's full name")
    if (not partial or "customerPhone" in payload) and not re.fullmatch(r"\+91 [6-9]\d{4} \d{5}", phone):
        raise HTTPException(status_code=400, detail="Phone must use the format +91 98765 43210")
    if not partial or "deliveryDate" in payload:
        try:
            handover = datetime.date.fromisoformat(delivery)
            if handover > datetime.date.today():
                raise ValueError
        except ValueError:
            raise HTTPException(status_code=400, detail="Handover date must be a valid past or current date")

# --- Authentication Routes ---

@app.post("/api/register")
def register(req: RegisterRequest, db: Session = Depends(get_db)):
    # Check if user already exists
    existing = db.query(User).filter(User.username == req.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    new_user = User(
        username=req.username,
        hashed_password=hash_password(req.password),
        role="pilot"
    )
    db.add(new_user)
    db.commit()
    return {"message": "User registered successfully"}

@app.post("/api/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username).first()
    if not user or not verify_password(req.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password"
        )
    
    # Generate JWT token
    token = create_access_token({"sub": user.username, "role": user.role})
    return {"token": token, "username": user.username, "role": user.role}

# --- Vehicle CRUD Routes ---

@app.get("/api/vehicles")
def list_vehicles(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Auto-migrate legacy database records if present
    db.query(Vehicle).filter(Vehicle.model == 'Comet').update({Vehicle.model: 'CT2'}, synchronize_session=False)
    db.query(Vehicle).filter(Vehicle.model == 'Cosmo').update({Vehicle.model: 'CO1'}, synchronize_session=False)
    db.commit()

    vehicles = db.query(Vehicle).all()
    return vehicles

@app.get("/api/vehicles/{vehicle_id}")
def get_vehicle(
    vehicle_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return vehicle

@app.post("/api/vehicles")
def create_vehicle(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    validate_vehicle_identity(payload)
    # Ensure unique VIN
    vin_val = payload.get("vin", "").strip().upper()
    if not vin_val:
        raise HTTPException(status_code=400, detail="VIN is required")
        
    existing = db.query(Vehicle).filter(Vehicle.vin == vin_val).first()
    if existing:
        raise HTTPException(status_code=400, detail="Vehicle with this VIN already registered")

    now_str = datetime.datetime.utcnow().isoformat() + "Z"
    
    # Initialize defaults
    services = payload.get("services") or [
        {"serviceNumber": 1, "dueKm": 1000, "completedKm": 0, "date": "", "technician": "", "issues": ""},
        {"serviceNumber": 2, "dueKm": 5000, "completedKm": 0, "date": "", "technician": "", "issues": ""},
        {"serviceNumber": 3, "dueKm": 10000, "completedKm": 0, "date": "", "technician": "", "issues": ""},
        {"serviceNumber": 4, "dueKm": 20000, "completedKm": 0, "date": "", "technician": "", "issues": ""}
    ]
    battery = payload.get("batteryReplacement") or {
        "affected": False,
        "campaignId": "",
        "status": "not_affected",
        "oldSerial": "",
        "newSerial": "",
        "replacementDate": "",
        "technician": "",
        "customerConfirmed": False
    }

    new_v = Vehicle(
        id=str(uuid.uuid4()),
        vin=vin_val,
        model=payload.get("model", "CT2"),
        chassisNo=payload.get("chassisNo", ""),
        motorNo=payload.get("motorNo", ""),
        controllerNo=payload.get("controllerNo", ""),
        batteryPackNo=payload.get("batteryPackNo", ""),
        manufacturingDate=payload.get("manufacturingDate", ""),
        customerName=payload.get("customerName", ""),
        customerPhone=payload.get("customerPhone", ""),
        customerLocation=payload.get("customerLocation", ""),
        deliveryDate=payload.get("deliveryDate", ""),
        warrantyExpiryDate=payload.get("warrantyExpiryDate", ""),
        contactHistory=payload.get("contactHistory") or [],
        issueCode=payload.get("issueCode", ""),
        issueReportedDate=payload.get("issueReportedDate", ""),
        currentKm=payload.get("currentKm", 0),
        registrationStatus=payload.get("registrationStatus", "delivered"),
        registrationNumber=payload.get("registrationNumber", ""),
        registrationDates=payload.get("registrationDates") or {"delivered": payload.get("deliveryDate", "")},
        registrationNotes=payload.get("registrationNotes") or {},
        batteryReplacement=battery,
        services=services,
        kmLog=payload.get("kmLog") or [],
        createdAt=now_str,
        updatedAt=now_str
    )
    db.add(new_v)
    db.commit()
    db.refresh(new_v)
    return new_v

@app.put("/api/vehicles/{vehicle_id}")
def update_vehicle(
    vehicle_id: str,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    validate_vehicle_identity(payload, partial=True)
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    # Update columns
    v.model = payload.get("model", v.model)
    v.chassisNo = payload.get("chassisNo", v.chassisNo)
    v.motorNo = payload.get("motorNo", v.motorNo)
    v.controllerNo = payload.get("controllerNo", v.controllerNo)
    v.batteryPackNo = payload.get("batteryPackNo", v.batteryPackNo)
    v.manufacturingDate = payload.get("manufacturingDate", v.manufacturingDate)
    
    v.customerName = payload.get("customerName", v.customerName)
    v.customerPhone = payload.get("customerPhone", v.customerPhone)
    v.customerLocation = payload.get("customerLocation", v.customerLocation)
    v.deliveryDate = payload.get("deliveryDate", v.deliveryDate)
    v.warrantyExpiryDate = payload.get("warrantyExpiryDate", v.warrantyExpiryDate)
    v.contactHistory = payload.get("contactHistory", v.contactHistory)
    v.issueCode = payload.get("issueCode", v.issueCode)
    v.issueReportedDate = payload.get("issueReportedDate", v.issueReportedDate)
    
    v.currentKm = payload.get("currentKm", v.currentKm)
    v.registrationStatus = payload.get("registrationStatus", v.registrationStatus)
    v.registrationNumber = payload.get("registrationNumber", v.registrationNumber)
    
    v.registrationDates = payload.get("registrationDates", v.registrationDates)
    v.registrationNotes = payload.get("registrationNotes", v.registrationNotes)
    v.batteryReplacement = payload.get("batteryReplacement", v.batteryReplacement)
    v.services = payload.get("services", v.services)
    v.kmLog = payload.get("kmLog", v.kmLog)
    
    v.updatedAt = datetime.datetime.utcnow().isoformat() + "Z"
    db.commit()
    db.refresh(v)
    return v

@app.delete("/api/vehicles/{vehicle_id}")
def delete_vehicle(
    vehicle_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    db.delete(v)
    db.commit()
    return {"message": "Vehicle deleted successfully"}

# --- Bulk Import Route ---

@app.post("/api/import")
async def import_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    contents = await file.read()
    filename = file.filename.lower()

    if filename.endswith(".xlsx"):
        import openpyxl
        buffer = io.BytesIO(contents)
        wb = openpyxl.load_workbook(buffer, data_only=True)
        ws = wb["EV Import Template"] if "EV Import Template" in wb.sheetnames else wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in r])
    else:
        buffer = io.StringIO(contents.decode('utf-8'))
        reader = csv.reader(buffer)
        rows = list(reader)

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="CSV is empty or missing data rows")

    headers = [h.strip() for h in rows[0]]
    if "vin" not in headers:
        raise HTTPException(status_code=400, detail="Missing required 'vin' column in CSV")

    vin_idx = headers.index("vin")
    now_str = datetime.datetime.utcnow().isoformat() + "Z"
    added = 0
    updated = 0

    for r in range(1, len(rows)):
        row = rows[r]
        if not row or len(row) < len(headers):
            if len(row) == 1 and row[0] == "":
                continue
            continue # skip incomplete rows

        vin_val = row[vin_idx].strip().upper()
        if not vin_val:
            continue # skip empty vin

        # Map row to dictionary
        vehicle_row = {}
        for idx, h in enumerate(headers):
            vehicle_row[h] = row[idx].strip() if idx < len(row) else ""

        validate_vehicle_identity({
            "vin": vin_val,
            "customerName": vehicle_row.get("customerName", ""),
            "customerPhone": vehicle_row.get("customerPhone", ""),
            "deliveryDate": vehicle_row.get("deliveryDate", "")
        })

        # Parse variables
        current_km = int(vehicle_row.get("currentKm", "0") or "0")
        model = vehicle_row.get("model", "CT2")
        mfg_date = vehicle_row.get("manufacturingDate") or datetime.date.today().isoformat()
        del_date = vehicle_row.get("deliveryDate") or datetime.date.today().isoformat()

        is_aff = (vehicle_row.get("batteryReplacementAffected", "")).lower()
        affected = (is_aff in ["true", "yes", "1"])

        is_conf = (vehicle_row.get("batteryReplacementCustomerConfirmed", "")).lower()
        confirmed = (is_conf in ["true", "yes", "1"])

        battery = {
            "affected": affected,
            "campaignId": vehicle_row.get("batteryReplacementCampaignId", ""),
            "status": vehicle_row.get("batteryReplacementStatus", "") or ("pending" if affected else "not_affected"),
            "oldSerial": vehicle_row.get("batteryReplacementOldSerial", ""),
            "newSerial": vehicle_row.get("batteryReplacementNewSerial", ""),
            "replacementDate": vehicle_row.get("batteryReplacementDate", ""),
            "technician": vehicle_row.get("batteryReplacementTechnician", ""),
            "customerConfirmed": confirmed,
            "reportedAt": vehicle_row.get("issueReportedDate", "")
        }

        last_contact = vehicle_row.get("lastCustomerContactDate", "")
        contact_history = [{
            "date": last_contact,
            "channel": "import",
            "outcome": vehicle_row.get("lastCustomerContactOutcome", "connected") or "connected",
            "note": "Imported latest customer contact",
        }] if last_contact else []

        reg_status = vehicle_row.get("registrationStatus", "delivered")
        reg_number = vehicle_row.get("registrationNumber", "")

        # Look up existing vehicle
        existing = db.query(Vehicle).filter(Vehicle.vin == vin_val).first()

        if existing:
            # Update
            existing.model = model
            existing.chassisNo = vin_val
            existing.motorNo = vehicle_row.get("motorNo") or existing.motorNo
            existing.controllerNo = vehicle_row.get("controllerNo") or existing.controllerNo
            existing.batteryPackNo = vehicle_row.get("batteryPackNo") or existing.batteryPackNo
            existing.manufacturingDate = mfg_date
            
            existing.customerName = vehicle_row.get("customerName") or existing.customerName
            existing.customerPhone = vehicle_row.get("customerPhone") or existing.customerPhone
            existing.customerLocation = vehicle_row.get("customerLocation") or existing.customerLocation
            existing.deliveryDate = del_date
            existing.warrantyExpiryDate = vehicle_row.get("warrantyExpiryDate") or existing.warrantyExpiryDate
            existing.contactHistory = contact_history or existing.contactHistory
            existing.issueCode = vehicle_row.get("issueCode") or existing.issueCode
            existing.issueReportedDate = vehicle_row.get("issueReportedDate") or existing.issueReportedDate

            if current_km != existing.currentKm:
                existing.currentKm = current_km
                # Update monthly log
                month_key = datetime.datetime.now().strftime("%Y-%m")
                km_log = list(existing.kmLog) if existing.kmLog else []
                # Remove duplicate months if already exists
                km_log = [k for k in km_log if k.get("month") != month_key]
                km_log.append({"month": month_key, "km": current_km})
                km_log.sort(key=lambda x: x["month"])
                existing.kmLog = km_log

            if reg_status != existing.registrationStatus or reg_number != existing.registrationNumber:
                existing.registrationStatus = reg_status
                existing.registrationNumber = reg_number
                reg_dates = dict(existing.registrationDates or {})
                if reg_status not in reg_dates:
                    reg_dates[reg_status] = datetime.date.today().isoformat()
                existing.registrationDates = reg_dates
                if reg_number:
                    reg_notes = dict(existing.registrationNotes or {})
                    reg_notes["completed"] = reg_number
                    existing.registrationNotes = reg_notes

            existing.batteryReplacement = battery
            existing.updatedAt = now_str
            updated += 1
        else:
            # Create
            new_v = Vehicle(
                id=str(uuid.uuid4()),
                vin=vin_val,
                model=model,
                chassisNo=vin_val,
                motorNo=vehicle_row.get("motorNo", ""),
                controllerNo=vehicle_row.get("controllerNo", ""),
                batteryPackNo=vehicle_row.get("batteryPackNo", ""),
                manufacturingDate=mfg_date,
                customerName=vehicle_row.get("customerName", ""),
                customerPhone=vehicle_row.get("customerPhone", ""),
                customerLocation=vehicle_row.get("customerLocation", ""),
                deliveryDate=del_date,
                warrantyExpiryDate=vehicle_row.get("warrantyExpiryDate", ""),
                contactHistory=contact_history,
                issueCode=vehicle_row.get("issueCode", ""),
                issueReportedDate=vehicle_row.get("issueReportedDate", ""),
                currentKm=current_km,
                kmLog=[{"month": del_date[:7], "km": current_km}] if current_km > 0 else [],
                registrationStatus=reg_status,
                registrationNumber=reg_number,
                registrationDates={
                    "delivered": del_date,
                    reg_status: datetime.date.today().isoformat()
                },
                registrationNotes={"completed": reg_number} if reg_number else {},
                services=[
                    {"serviceNumber": 1, "dueKm": 1000, "completedKm": 0, "date": "", "technician": "", "issues": ""},
                    {"serviceNumber": 2, "dueKm": 5000, "completedKm": 0, "date": "", "technician": "", "issues": ""},
                    {"serviceNumber": 3, "dueKm": 10000, "completedKm": 0, "date": "", "technician": "", "issues": ""},
                    {"serviceNumber": 4, "dueKm": 20000, "completedKm": 0, "date": "", "technician": "", "issues": ""}
                ],
                batteryReplacement=battery,
                createdAt=now_str,
                updatedAt=now_str
            )
            db.add(new_v)
            added += 1

    db.commit()
    return {"type": "csv", "added": added, "updated": updated}


# --- AI Action Centre Routes ---

def action_centre_payload(db: Session):
    vehicles = db.query(Vehicle).all()
    cases = sync_cases(db, vehicles)
    vehicle_map = {vehicle.id: vehicle for vehicle in vehicles}
    active_cases = db.query(ActionCase).filter(ActionCase.status.notin_(["resolved", "auto_closed"])).all()
    counts = {priority: sum(case.priority == priority for case in active_cases) for priority in ("critical", "high", "medium")}
    now = iso_now()
    recent_events = db.query(ActionAuditEvent).order_by(ActionAuditEvent.createdAt.desc()).limit(12).all()
    return {
        "cases": [serialize_case(case, vehicle_map) for case in cases],
        "summary": {
            **counts,
            "active": len(active_cases),
            "overdueSla": sum(case.slaDeadline < now for case in active_cases),
        },
        "audit": [{"id":event.id, "caseId":event.caseId, "actor":event.actor, "eventType":event.eventType, "details":event.details or {}, "createdAt":event.createdAt} for event in recent_events],
        "generatedAt": now,
        "engineVersion": "action-centre-v1",
    }


@app.get("/api/action-centre")
def get_action_centre(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    return action_centre_payload(db)


@app.post("/api/action-centre/enrich")
def enrich_action_queue(
    request: EnrichRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    vehicles = db.query(Vehicle).all()
    cases = sync_cases(db, vehicles)
    enrichment = enrich_action_centre(db, cases, force=request.force)
    payload = action_centre_payload(db)
    payload["brief"] = enrichment["brief"]
    payload["briefSource"] = enrichment["source"]
    payload["queueFingerprint"] = enrichment["queueFingerprint"]
    return payload


@app.post("/api/action-cases/{case_id}/draft-whatsapp")
def create_whatsapp_draft(
    case_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    case = db.query(ActionCase).filter(ActionCase.id == case_id).first()
    if not case or not case.vehicleIds:
        raise HTTPException(status_code=404, detail="Action case not found")
    vehicle = db.query(Vehicle).filter(Vehicle.id == case.vehicleIds[0]).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Affected vehicle not found")
    return draft_whatsapp(case, vehicle)


def add_action_audit(db, case, actor, event_type, details):
    db.add(ActionAuditEvent(
        id=str(uuid.uuid4()), caseId=case.id, actor=actor,
        eventType=event_type, details=details, createdAt=iso_now()
    ))


@app.post("/api/action-cases/{case_id}/actions")
def perform_case_action(
    case_id: str,
    request: ActionRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user.get("role") != "master" and current_user.get("sub") != "ismailadmin":
        raise HTTPException(status_code=403, detail="Only ismailadmin or master may approve operational actions")
    case = db.query(ActionCase).filter(ActionCase.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="Action case not found")
    if case.status in ("resolved", "auto_closed"):
        raise HTTPException(status_code=409, detail="This case is already closed")

    action_type = request.type
    payload = request.payload or {}
    actor = current_user.get("sub", "unknown")
    now = iso_now()
    linked_vehicles = db.query(Vehicle).filter(Vehicle.id.in_(case.vehicleIds or [])).all()

    if action_type == "assign_technician":
        owner = str(payload.get("owner", "")).strip()
        if not owner:
            raise HTTPException(status_code=400, detail="Select a technician")
        case.assignedOwner = owner
        case.status = "assigned"
        case.assignedAt = now
        case.actionedAt = now
        for vehicle in linked_vehicles:
            if case.caseType == "battery_recall":
                battery = dict(vehicle.batteryReplacement or {})
                battery["technician"] = owner
                vehicle.batteryReplacement = battery
        details = {"owner": owner}

    elif action_type == "create_campaign":
        name = str(payload.get("name", "")).strip()
        region = str(payload.get("region", "")).strip()
        owner = str(payload.get("owner", "")).strip()
        if not all((name, region, owner)):
            raise HTTPException(status_code=400, detail="Campaign name, region and owner are required")
        campaign = ServiceCampaign(
            id=f"SC-{uuid.uuid4().hex[:12].upper()}",
            name=name, region=region,
            issueCode=next((vehicle.issueCode for vehicle in linked_vehicles if vehicle.issueCode), case.caseType),
            vehicleIds=[vehicle.id for vehicle in linked_vehicles], owner=owner,
            status="planned", createdAt=now,
        )
        db.add(campaign)
        case.assignedOwner = owner
        case.status = "in_progress"
        case.assignedAt = case.assignedAt or now
        case.actionedAt = now
        details = {"campaignId": campaign.id, "name": name, "region": region, "owner": owner, "vehicleCount": len(linked_vehicles)}

    elif action_type == "escalate":
        owner = str(payload.get("owner", "")).strip()
        note = str(payload.get("note", "")).strip()
        if not owner or not note:
            raise HTTPException(status_code=400, detail="Escalation owner and note are required")
        case.assignedOwner = owner
        case.status = "escalated"
        case.assignedAt = case.assignedAt or now
        case.actionedAt = now
        details = {"owner": owner, "note": note}

    elif action_type == "resolve":
        note = str(payload.get("note", "")).strip()
        if not note:
            raise HTTPException(status_code=400, detail="A resolution note is required")
        case.status = "resolved"
        case.actionedAt = now
        case.resolvedAt = now
        details = {"note": note}

    elif action_type == "log_contact":
        vehicle_id = str(payload.get("vehicleId", "")).strip()
        vehicle = next((item for item in linked_vehicles if item.id == vehicle_id), None)
        if not vehicle:
            raise HTTPException(status_code=400, detail="Select an affected vehicle")
        outcome = str(payload.get("outcome", "opened")).strip()
        channel = str(payload.get("channel", "whatsapp")).strip()
        contacts = list(vehicle.contactHistory or [])
        contacts.append({"date":datetime.date.today().isoformat(), "channel":channel, "outcome":outcome, "note":str(payload.get("note", "")).strip()})
        vehicle.contactHistory = contacts
        if case.status == "open":
            case.status = "in_progress"
        case.actionedAt = now
        details = {"vehicleId":vehicle_id, "channel":channel, "outcome":outcome}

    else:
        raise HTTPException(status_code=400, detail="Unsupported action type")

    case.updatedAt = now
    add_action_audit(db, case, actor, action_type, details)
    db.commit()
    return action_centre_payload(db)

# --- Session Tracking Routes ---
from api.database import SessionLog
from fastapi import Request

@app.post("/api/tracking/session")
def start_tracking_session(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub", "unknown")
    
    # Extract IP address
    x_forwarded_for = request.headers.get("x-forwarded-for")
    if x_forwarded_for:
        ip_address = x_forwarded_for.split(",")[0].strip()
    else:
        ip_address = request.headers.get("x-real-ip") or (request.client.host if request.client else "127.0.0.1")

    # Extract Location from Vercel Geolocation headers
    city = request.headers.get("x-vercel-ip-city")
    region = request.headers.get("x-vercel-ip-country-region")
    country = request.headers.get("x-vercel-ip-country")
    
    if city or region or country:
        parts = [p for p in [city, region, country] if p]
        location = ", ".join(parts)
    else:
        # Fallback for local testing
        if ip_address in ("127.0.0.1", "::1"):
            location = "Localhost Development"
        else:
            location = "Local Network / Private IP"

    now_str = datetime.datetime.utcnow().isoformat() + "Z"
    session_id = str(uuid.uuid4())
    
    new_session = SessionLog(
        id=session_id,
        username=username,
        ipAddress=ip_address,
        location=location,
        startedAt=now_str,
        lastHeartbeat=now_str,
        durationSeconds=0
    )
    db.add(new_session)
    db.commit()
    
    return {"session_id": session_id, "location": location, "ip": ip_address}


@app.post("/api/tracking/heartbeat/{session_id}")
def tracking_heartbeat(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    session_record = db.query(SessionLog).filter(SessionLog.id == session_id).first()
    if not session_record:
        raise HTTPException(status_code=404, detail="Tracking session not found")
        
    now = datetime.datetime.utcnow()
    # Calculate duration
    try:
        started_at = datetime.datetime.fromisoformat(session_record.startedAt.replace("Z", "+00:00"))
        started_naive = started_at.replace(tzinfo=None)
        elapsed = now - started_naive
        session_record.durationSeconds = max(0, int(elapsed.total_seconds()))
    except Exception:
        session_record.durationSeconds += 30

    session_record.lastHeartbeat = now.isoformat() + "Z"
    db.commit()
    
    return {"status": "ok", "duration": session_record.durationSeconds}


@app.get("/api/tracking/stats")
def get_tracking_stats(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Check if the user is authorized (role must be master)
    user_role = current_user.get("role", "pilot")
    if user_role != "master":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access Denied: Only Master Admin can view usage analytics."
        )

    # 1. Total site opens
    total_opens = db.query(SessionLog).count()

    # 2. Active sessions (heartbeat within the last 90 seconds to allow buffer)
    now = datetime.datetime.utcnow()
    cutoff_time = now - datetime.timedelta(seconds=90)
    cutoff_str = cutoff_time.isoformat() + "Z"
    
    active_sessions = db.query(SessionLog).filter(SessionLog.lastHeartbeat >= cutoff_str).count()

    # 3. Total usage duration in hours
    total_seconds = db.query(__import__('sqlalchemy').func.sum(SessionLog.durationSeconds)).scalar() or 0
    total_hours = round(total_seconds / 3600.0, 2)

    # 4. Location summary
    locations_query = db.query(SessionLog.location, __import__('sqlalchemy').func.count(SessionLog.id)).group_by(SessionLog.location).all()
    locations = [{"location": loc or "Unknown", "count": count} for loc, count in locations_query]

    # 5. User usage ranks
    users_query = db.query(
        SessionLog.username,
        __import__('sqlalchemy').func.count(SessionLog.id),
        __import__('sqlalchemy').func.sum(SessionLog.durationSeconds)
    ).group_by(SessionLog.username).all()
    
    user_activity = [
        {"username": user, "opens": count, "durationMinutes": round((duration or 0) / 60.0, 1)}
        for user, count, duration in users_query
    ]

    # 6. Recent sessions (last 20 logs)
    recent_query = db.query(SessionLog).order_by(SessionLog.startedAt.desc()).limit(20).all()
    recent_sessions = [
        {
            "id": s.id,
            "username": s.username,
            "ipAddress": s.ipAddress,
            "location": s.location,
            "startedAt": s.startedAt,
            "durationSeconds": s.durationSeconds
        }
        for s in recent_query
    ]

    return {
        "totalOpens": total_opens,
        "activeSessions": active_sessions,
        "totalUsageHours": total_hours,
        "locations": locations,
        "userActivity": user_activity,
        "recentSessions": recent_sessions
    }

# --- Static File Serving (For local testing fallback) ---

# Mount static CSS & JS
if os.path.exists("css"):
    app.mount("/css", StaticFiles(directory="css"), name="css")
if os.path.exists("js"):
    app.mount("/js", StaticFiles(directory="js"), name="js")

@app.get("/")
def serve_home():
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    return JSONResponse({"message": "PulseEV Dashboard API is online. Static files missing."})

@app.get("/{path:path}")
def serve_wildcard(path: str):
    # If the file path exists directly, return it (e.g. template files)
    if os.path.exists(path) and os.path.isfile(path):
        return FileResponse(path)
    # Redirect to index.html for SPA router fallbacks
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    return JSONResponse({"status": "not_found"}, status_code=404)
