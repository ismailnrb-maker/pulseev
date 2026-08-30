#!/usr/bin/env python
import csv
import openpyxl
import datetime
import random
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def generate_vehicles_data(count=200):
    random.seed(42)
    
    first_names = ['Rajesh', 'Priya', 'Amit', 'Sneha', 'Vikash', 'Ananya', 'Deepak', 'Kavita', 'Rohit', 'Sunita', 'Aarav', 'Vihaan', 'Aditya', 'Sai', 'Arjun', 'Krishna', 'Ishaan', 'Shaurya', 'Pranav', 'Aryan', 'Diya', 'Ananya', 'Aanya', 'Pihu', 'Prisha', 'Saanvi', 'Anika', 'Zara', 'Meera', 'Riya', 'Rahul', 'Sanjay', 'Manoj', 'Rohan', 'Karan', 'Dev', 'Vijay', 'Raj', 'Alok', 'Vikram']
    last_names = ['Mehra', 'Sharma', 'Joshi', 'Kulkarni', 'Gupta', 'Reddy', 'Nair', 'Singh', 'Deshmukh', 'Patil', 'Kumar', 'Verma', 'Yadav', 'Patel', 'Das', 'Choudhury', 'Banerjee', 'Mishra', 'Trivedi', 'Rao', 'Bose', 'Pillai', 'Jha', 'Kapoor', 'Mehta', 'Grover', 'Sen', 'Dutta', 'Chatterjee']
    
    cities = ['Mumbai, MH', 'Delhi, DL', 'Bengaluru, KA', 'Pune, MH', 'Hyderabad, TS', 'Chennai, TN', 'Ahmedabad, GJ', 'Jaipur, RJ', 'Kolkata, WB', 'Lucknow, UP']
    technicians = ['Vikram Singh', 'Rajesh Kumar', 'Arjun Patel', 'Sanjay Verma', 'Manoj Sharma', 'Amit Yadav']
    
    # State mapping for registration numbers
    state_mapping = {
        'Mumbai, MH': 'MH-02', 'Pune, MH': 'MH-12', 'Delhi, DL': 'DL-3C', 
        'Bengaluru, KA': 'KA-03', 'Hyderabad, TS': 'TS-09', 'Chennai, TN': 'TN-09',
        'Ahmedabad, GJ': 'GJ-01', 'Jaipur, RJ': 'RJ-14', 'Kolkata, WB': 'WB-02',
        'Lucknow, UP': 'UP-32'
    }
    
    vehicles = []
    
    # Keep models saturated at 70% Comet, 30% Cosmo
    models = ['Comet'] * 140 + ['Cosmo'] * 60
    
    # Keep battery at 50% affected (100 vehicles)
    # Among affected (100 vehicles), 60% replaced (60 vehicles), 20% in_progress (20), 20% pending (20)
    battery_statuses = (
        [('completed', True)] * 60 +
        [('in_progress', False)] * 20 +
        [('pending', False)] * 20 +
        [('not_affected', False)] * 100
    )
    
    # Shuffle models and battery statuses independently but with fixed seeds so they are deterministic
    random.shuffle(models)
    random.shuffle(battery_statuses)
    
    # Generate VINs from MAT45678901234101 to MAT45678901234300 (exactly 200 unique 17-char VINs)
    vins = [f"MAT45678901234{idx}" for idx in range(101, 301)]
    
    # Dates from 2023 to 2026
    start_date = datetime.date(2023, 1, 1)
    end_date = datetime.date(2026, 8, 25) # Capped at August 2026
    days_range = (end_date - start_date).days
    
    for i in range(count):
        vin = vins[i]
        model = models[i]
        
        # Odometer mileage
        # Choose a random delivery date
        random_days = random.randint(0, days_range)
        mfg_date = start_date + datetime.timedelta(days=random_days)
        # delivery date is 5 to 25 days after mfg date
        del_date = mfg_date + datetime.timedelta(days=random.randint(5, 25))
        
        # If dates are in the future relative to 2026-08-30, clamp them
        if mfg_date > datetime.date(2026, 8, 30):
            mfg_date = datetime.date(2026, 8, 30) - datetime.timedelta(days=random.randint(5, 20))
        if del_date > datetime.date(2026, 8, 30):
            del_date = datetime.date(2026, 8, 30)
            
        mfg_date_str = mfg_date.isoformat()
        del_date_str = del_date.isoformat()
        
        # Calculate months active
        months_active = (datetime.date(2026, 8, 30) - del_date).days // 30
        months_active = max(1, months_active)
        
        # Current km based on months active
        km_per_month = random.randint(600, 1200)
        current_km = months_active * km_per_month
        
        # Customer Location
        location = random.choice(cities)
        
        # Customer info
        cust_name = f"{random.choice(first_names)} {random.choice(last_names)}"
        cust_phone = f"+91-98765{random.randint(10000, 99999)}"
        
        # Registration Status
        # If delivered > 2 months, completion is high (completed)
        if months_active >= 2:
            reg_status = 'completed'
            state_prefix = state_mapping[location]
            letters = "".join(random.choices("ABCDEFGHIJKLMNOPQRSTUVWXYZ", k=2))
            nums = f"{random.randint(1000, 9999)}"
            reg_number = f"{state_prefix}-{letters}-{nums}"
        else:
            reg_status = random.choice(['completed', 'submitted', 'documents_pending', 'delivered'])
            if reg_status == 'completed':
                state_prefix = state_mapping[location]
                letters = "".join(random.choices("ABCDEFGHIJKLMNOPQRSTUVWXYZ", k=2))
                nums = f"{random.randint(1000, 9999)}"
                reg_number = f"{state_prefix}-{letters}-{nums}"
            else:
                reg_number = ''
                
        # Services
        services = []
        due_milestones = [1000, 5000, 10000, 20000]
        for service_idx, due_km in enumerate(due_milestones, start=1):
            service_completed = current_km >= due_km
            if service_completed:
                completed_km = due_km + random.randint(-200, 500)
                completed_km = max(100, completed_km)
                days_to_reach = int((completed_km / km_per_month) * 30)
                service_date = del_date + datetime.timedelta(days=days_to_reach)
                if service_date > datetime.date(2026, 8, 30):
                    service_date = datetime.date(2026, 8, 30) - datetime.timedelta(days=random.randint(1, 15))
                service_date_str = service_date.isoformat()
                technician = random.choice(technicians)
                issues = random.choice(['None', 'None', 'None', 'Wheel alignment check', 'Minor brake pad adjustment', 'Software patch applied'])
            else:
                completed_km = 0
                service_date_str = ''
                technician = ''
                issues = ''
                
            services.append({
                "serviceNumber": service_idx,
                "dueKm": due_km,
                "completedKm": completed_km,
                "date": service_date_str,
                "technician": technician,
                "issues": issues
            })
            
        # Odometer kmLog
        km_log = []
        for s in services:
            if s["completedKm"] > 0:
                log_month = s["date"][:7]
                km_log.append({"month": log_month, "km": s["completedKm"]})
        current_month = "2026-08"
        if not any(k["month"] == current_month for k in km_log):
            km_log.append({"month": current_month, "km": current_km})
        km_log.sort(key=lambda x: x["month"])
        unique_km_log = []
        seen_months = set()
        for k in km_log:
            if k["month"] not in seen_months:
                unique_km_log.append(k)
                seen_months.add(k["month"])
        km_log = unique_km_log
        
        # Battery Recall Info
        bat_status, cust_conf = battery_statuses[i]
        is_affected = bat_status != 'not_affected'
        campaign_id = 'BC-2024-001' if is_affected else ''
        
        battery_prefix = 'BP-LFP-96' if model == 'Comet' else 'BP-NMC-72'
        old_serial = f"{battery_prefix}{i+1:03d}"
        
        if bat_status == 'completed':
            new_serial = f"{old_serial}-R"
            replace_days = random.randint(90, 360)
            replace_date = del_date + datetime.timedelta(days=replace_days)
            if replace_date > datetime.date(2026, 8, 30):
                replace_date = datetime.date(2026, 8, 30) - datetime.timedelta(days=random.randint(10, 60))
            replace_date_str = replace_date.isoformat()
            technician = random.choice(technicians)
        else:
            new_serial = ''
            replace_date_str = ''
            if bat_status == 'in_progress':
                technician = random.choice(technicians)
            else:
                technician = ''
                
        battery_replacement = {
            "affected": is_affected,
            "campaignId": campaign_id,
            "status": bat_status,
            "oldSerial": old_serial if is_affected else '',
            "newSerial": new_serial,
            "replacementDate": replace_date_str,
            "technician": technician,
            "customerConfirmed": cust_conf
        }
        
        chassis_no = f"CH-{del_date.year}-{i+1:03d}"
        motor_no = f"MT-ZF-78{i+1:03d}"
        controller_no = f"CT-INV-44{i+1:03d}"
        
        vehicles.append({
            "vin": vin,
            "model": model,
            "chassisNo": chassis_no,
            "motorNo": motor_no,
            "controllerNo": controller_no,
            "batteryPackNo": old_serial,
            "manufacturingDate": mfg_date_str,
            "customerName": cust_name,
            "customerPhone": cust_phone,
            "customerLocation": location,
            "deliveryDate": del_date_str,
            "currentKm": current_km,
            "registrationStatus": reg_status,
            "registrationNumber": reg_number,
            "batteryReplacement": battery_replacement,
            "services": services,
            "kmLog": km_log
        })
        
    return vehicles

def create_excel_template():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # TAB 1: Instructions & Reference
    # -------------------------------------------------------------
    ws_ref = wb.active
    ws_ref.title = "Instructions & Reference"
    ws_ref.views.sheetView[0].showGridLines = True
    
    # Colors
    navy_fill = PatternFill(start_color="0C1224", end_color="0C1224", fill_type="solid")
    light_blue_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    zebra_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    # Fonts
    title_font = Font(name="Segoe UI", size=16, bold=True, color="0C1224")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1E3A8A")
    header_font_white = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    normal_font = Font(name="Segoe UI", size=10, color="334155")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    italic_font = Font(name="Segoe UI", size=9, color="64748B")
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Sheet Title
    ws_ref.cell(row=2, column=2, value="PulseEV — Centralized EV Lifecycle Intelligence").font = title_font
    ws_ref.cell(row=3, column=2, value="Bulk Upload & Import Instructions").font = Font(name="Segoe UI", size=11, italic=True, color="475569")
    
    # General Steps Section
    ws_ref.cell(row=5, column=2, value="HOW TO USE THIS TEMPLATE:").font = section_font
    instructions = [
        "1. Complete the 'EV Import Template' tab with your vehicle records.",
        "2. To UPDATE an existing EV profile: Ensure the 'vin' matches an existing vehicle in PulseEV.",
        "3. To ADD a new EV profile: Input a new 17-digit VIN. Other fields will be initialized to their defaults if empty.",
        "4. Column headers must remain exactly as named. Do not delete or rename any headers.",
        "5. Save this Excel sheet, or export/save the template tab as a CSV file.",
        "6. In the PulseEV dashboard, click the 'Import' button in the header actions, and select your file."
    ]
    for idx, inst in enumerate(instructions):
        cell = ws_ref.cell(row=6 + idx, column=2, value=inst)
        cell.font = normal_font
        cell.alignment = left_align

    # Field Reference Header
    ws_ref.cell(row=14, column=2, value="FIELD DATA DICTIONARY:").font = section_font
    
    ref_headers = ["Field (Column)", "Required?", "Accepted Values / Format", "Description & Rules"]
    for col_idx, h in enumerate(ref_headers, start=2):
        cell = ws_ref.cell(row=15, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws_ref.row_dimensions[15].height = 24

    field_rules = [
        ("vin", "YES", "Alphanumeric string", "VIN / Chassis Number (Primary Key)."),
        ("model", "YES", "Text (Comet or Cosmo)", "Specific EV model name. Must match allowed list."),
        ("motorNo", "YES", "Alphanumeric string (e.g. MT-ZF-78001)", "Electric Motor serial identification number."),
        ("controllerNo", "NO", "Alphanumeric string (e.g. CT-INV-44001)", "Controller unit serial number."),
        ("batteryPackNo", "NO", "Alphanumeric string (e.g. BP-LFP-96001)", "Active Battery Pack serial number."),
        ("manufacturingDate", "YES", "YYYY-MM-DD (Date)", "Date of purchase."),
        ("customerName", "YES", "Full Name string", "Assigned customer owner's name."),
        ("customerPhone", "YES", "Phone pattern (e.g. +91-98765-43210)", "Contact mobile/WhatsApp number."),
        ("customerLocation", "YES", "City, State (e.g. Mumbai, MH)", "Delivery region / operational city location."),
        ("deliveryDate", "YES", "YYYY-MM-DD (Date)", "Date of vehicle handover to customer."),
        ("currentKm", "NO", "Positive integer (e.g. 1500)", "Current odometer reading in kilometers (Defaults to 0)."),
        ("registrationStatus", "NO", "delivered / documents_pending / submitted / completed", "Current status of RTO registration."),
        ("registrationNumber", "NO", "Registration plate string (e.g. MH-02-XX-1234)", "Assigned RTO vehicle registration plate number."),
        ("batteryReplacementAffected", "NO", "TRUE / FALSE", "Whether vehicle is part of a battery replacement campaign."),
        ("batteryReplacementCampaignId", "NO", "Alphanumeric string (e.g. BC-2024-001)", "Recall or upgrade campaign reference ID."),
        ("batteryReplacementStatus", "NO", "not_affected / pending / in_progress / completed", "Upgrade campaign implementation status."),
        ("batteryReplacementOldSerial", "NO", "Alphanumeric string", "Serial number of the decommissioned battery."),
        ("batteryReplacementNewSerial", "NO", "Alphanumeric string", "Serial number of the newly fitted battery pack."),
        ("batteryReplacementDate", "NO", "YYYY-MM-DD (Date)", "Date of battery service upgrade completion."),
        ("batteryReplacementTechnician", "NO", "Technician Name string", "Technician who carried out the replacement."),
        ("batteryReplacementCustomerConfirmed", "NO", "TRUE / FALSE", "Customer signed-off/confirmed the upgrade.")
    ]

    for row_idx, rule in enumerate(field_rules, start=16):
        is_zebra = (row_idx % 2 == 0)
        for col_idx, val in enumerate(rule, start=2):
            cell = ws_ref.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx == 2:
                cell.font = bold_font
            if col_idx == 3:
                cell.alignment = center_align
                if val == "YES":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
            if is_zebra:
                cell.fill = zebra_fill
        ws_ref.row_dimensions[row_idx].height = 20

    for col in ws_ref.columns:
        if col[0].column < 2 or col[0].column > 5:
            continue
        max_len = 0
        for cell in col:
            if cell.row >= 15:
                max_len = max(max_len, len(str(cell.value or '')))
        col_letter = get_column_letter(col[0].column)
        ws_ref.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws_ref.column_dimensions['A'].width = 3
    ws_ref.column_dimensions['B'].width = 35
    ws_ref.column_dimensions['C'].width = 12
    ws_ref.column_dimensions['D'].width = 45
    ws_ref.column_dimensions['E'].width = 65

    # -------------------------------------------------------------
    # TAB 2: EV Import Template
    # -------------------------------------------------------------
    ws_tpl = wb.create_sheet(title="EV Import Template")
    ws_tpl.views.sheetView[0].showGridLines = True
    
    headers = [
        "vin", "model", "motorNo", "controllerNo", "batteryPackNo",
        "manufacturingDate", "customerName", "customerPhone", "customerLocation",
        "deliveryDate", "currentKm", "registrationStatus", "registrationNumber",
        "batteryReplacementAffected", "batteryReplacementCampaignId", "batteryReplacementStatus",
        "batteryReplacementOldSerial", "batteryReplacementNewSerial", "batteryReplacementDate",
        "batteryReplacementTechnician", "batteryReplacementCustomerConfirmed"
    ]
    
    # Styled headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_tpl.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws_tpl.row_dimensions[1].height = 28
    
    # Generate the 200 customer data points
    vehicles_data = generate_vehicles_data(200)
    
    # Write the 200 records to the template
    for idx, v in enumerate(vehicles_data):
        row_idx = idx + 2
        is_zebra = (row_idx % 2 == 0)
        
        br = v["batteryReplacement"]
        row_values = [
            v["vin"],
            v["model"],
            v["motorNo"],
            v["controllerNo"],
            v["batteryPackNo"],
            v["manufacturingDate"],
            v["customerName"],
            v["customerPhone"],
            v["customerLocation"],
            v["deliveryDate"],
            v["currentKm"],
            v["registrationStatus"],
            v["registrationNumber"],
            "TRUE" if br["affected"] else "FALSE",
            br["campaignId"],
            br["status"],
            br["oldSerial"],
            br["newSerial"],
            br["replacementDate"],
            br["technician"],
            "TRUE" if br["customerConfirmed"] else "FALSE"
        ]
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws_tpl.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = left_align
            cell.border = thin_border
            if is_zebra:
                cell.fill = zebra_fill
        ws_tpl.row_dimensions[row_idx].height = 20

    # Auto sizing columns for template sheet based on headers and data
    for col in ws_tpl.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_tpl.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Data Validations (Dropdown limits)
    dv_model = DataValidation(type="list", formula1='"Comet,Cosmo"', allow_blank=True)
    dv_model.error = 'Your entry is not in the list of allowed vehicle models (Comet, Cosmo)'
    dv_model.errorTitle = 'Invalid Model'
    dv_model.prompt = 'Please select Comet or Cosmo'
    dv_model.promptTitle = 'Select EV Model'
    ws_tpl.add_data_validation(dv_model)
    dv_model.add("B2:B300")

    dv_reg = DataValidation(type="list", formula1='"delivered,documents_pending,submitted,completed"', allow_blank=True)
    dv_reg.error = 'Must choose delivered, documents_pending, submitted, or completed'
    dv_reg.errorTitle = 'Invalid Registration Status'
    dv_reg.prompt = 'Select registration stage'
    dv_reg.promptTitle = 'Select Status'
    ws_tpl.add_data_validation(dv_reg)
    dv_reg.add("L2:L300")

    dv_bat_status = DataValidation(type="list", formula1='"not_affected,pending,in_progress,completed"', allow_blank=True)
    dv_bat_status.error = 'Must choose not_affected, pending, in_progress, or completed'
    dv_bat_status.errorTitle = 'Invalid Upgrade Status'
    ws_tpl.add_data_validation(dv_bat_status)
    dv_bat_status.add("P2:P300")

    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_bool.error = 'Must enter TRUE or FALSE'
    dv_bool.errorTitle = 'Invalid Boolean Value'
    ws_tpl.add_data_validation(dv_bool)
    dv_bool.add("N2:N300")
    dv_bool.add("U2:U300")

    wb.save("ev_lifecycle_template.xlsx")
    print("Excel template 'ev_lifecycle_template.xlsx' (with 200 data points) created successfully.")

def create_csv_template():
    headers = [
        "vin", "model", "motorNo", "controllerNo", "batteryPackNo",
        "manufacturingDate", "customerName", "customerPhone", "customerLocation",
        "deliveryDate", "currentKm", "registrationStatus", "registrationNumber",
        "batteryReplacementAffected", "batteryReplacementCampaignId", "batteryReplacementStatus",
        "batteryReplacementOldSerial", "batteryReplacementNewSerial", "batteryReplacementDate",
        "batteryReplacementTechnician", "batteryReplacementCustomerConfirmed"
    ]
    
    vehicles_data = generate_vehicles_data(200)
    
    with open("ev_lifecycle_template.csv", mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for v in vehicles_data:
            br = v["batteryReplacement"]
            row_values = [
                v["vin"],
                v["model"],
                v["motorNo"],
                v["controllerNo"],
                v["batteryPackNo"],
                v["manufacturingDate"],
                v["customerName"],
                v["customerPhone"],
                v["customerLocation"],
                v["deliveryDate"],
                v["currentKm"],
                v["registrationStatus"],
                v["registrationNumber"],
                "TRUE" if br["affected"] else "FALSE",
                br["campaignId"],
                br["status"],
                br["oldSerial"],
                br["newSerial"],
                br["replacementDate"],
                br["technician"],
                "TRUE" if br["customerConfirmed"] else "FALSE"
            ]
            writer.writerow(row_values)
            
    print("CSV template 'ev_lifecycle_template.csv' (with 200 data points) created successfully.")

if __name__ == "__main__":
    create_excel_template()
    create_csv_template()
